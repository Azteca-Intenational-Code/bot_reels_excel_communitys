import subprocess

import uvicorn
from config import SessionLocal
from sqlalchemy import bindparam, text as sql_text
from GptAPi import GPT
from openpyxl import load_workbook
from calendar import month_name
from datetime import datetime, time, timedelta
import shutil
import random
import os
import sys
import io
def run_bot():

    # Configuración
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    plantilla_path = os.path.join(BASE_DIR, "Archivo de prueba reels para Niyi.xlsx")
    fecha_actual = datetime.today()
    carpeta_destino = os.path.join(BASE_DIR, "excel_campañas")
    mes_actual = datetime.today().month
    nombre_mes_actual = month_name[mes_actual]

    if not os.path.exists(carpeta_destino):
        os.makedirs(carpeta_destino)

    with SessionLocal() as session:
        result = session.execute(sql_text("SELECT campaign, commercial_services, residential_services, language FROM campaign"))
        rows = result.fetchall()

    campañas_deseadas = ["elite chicago spa"]
    # campañas_no_deseadas = ["botanica san gregorio", "osceola fence corporation", "elite chicago spa", "quick cleaning"]
    
    rows = [row for row in rows if row[0].strip().lower() in [c.lower() for c in campañas_deseadas]]
    
    # === IGNORAR CAMPAÑAS
    # rows = [row for row in rows if row[0].strip().lower() not in campañas_no_deseadas]

    if not rows:
        print("⚠️ La campaña 'Botanica San Gregorio' no fue encontrada en la base de datos.")
        exit()

    dias_semana = {
        0: "lunes", 1: "martes", 2: "miércoles", 3: "jueves",
        4: "viernes", 5: "sábado", 6: "domingo"
    }

    for row in rows:
        campaign_name, commercial_services, residential_services, lang = row
        campaign_key = campaign_name.lower().replace(" ", "_")
        print(f"\n📢 Procesando campaña: {campaign_name} [{campaign_key}]")

        excel_path = os.path.join(carpeta_destino, f"Reels_{campaign_key}.xlsx")
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
            ws = wb.active
        else:
            shutil.copy(plantilla_path, excel_path)
            wb = load_workbook(excel_path)
            ws = wb.active

        headers = [cell.value for cell in ws[1]]
        for col in ["Picture Url 1", "Time", "Facebook Title"]:
            if col not in headers:
                headers.append(col)
                ws.cell(row=1, column=len(headers)).value = col
        
        post_generados = 0
        idx_text = headers.index("Text") + 1
        idx_date = headers.index("Date") + 1
        idx_title = headers.index("Document title") + 1
        idx_yt_title = headers.index("Youtube Video Title") + 1
        idx_yt_tags = headers.index("Youtube Video Tags") + 1
        idx_fb_title = headers.index("Facebook Title") + 1
        idx_comment = headers.index("First Comment Text") + 1
        idx_tiktok = headers.index("TikTok Title") + 1
        idx_picture = headers.index("Picture Url 1") + 1
        idx_time = headers.index("Time") + 1
        idx_brand = headers.index("Brand name") + 1 if "Brand name" in headers else None

        cols_autoincrement = [
            "Document title", "Youtube Video Title", "Youtube Video Tags",
            "Facebook Title", "TikTok Title"
        ]

        idx_autoincrement = {
            col: headers.index(col) + 1 for col in cols_autoincrement if col in headers
        }

        with SessionLocal() as session:
            # Obtener ID numérico real de la campaña
            result = session.execute(sql_text("""
                SELECT id FROM campaign WHERE LOWER(TRIM(campaign)) = :nombre
            """), {"nombre": campaign_name.lower().strip()})
            row = result.fetchone()
            if row:
                id_campania = row[0]
            else:
                print(f"❌ No se encontró el ID para la campaña '{campaign_name}'")
                exit()

            # Usar el ID numérico para consultar contenido del día actual
            result = session.execute(sql_text("""
                SELECT id, dia, canal, servicio, tipo, descripcion, hashtags, sonido
                FROM contenido_semanal
                WHERE id_campaign = :camp
                ORDER BY
                    CASE LOWER(dia)
                        WHEN 'lunes' THEN 1
                        WHEN 'martes' THEN 2
                        WHEN 'miércoles' THEN 3
                        WHEN 'jueves' THEN 4
                        WHEN 'viernes' THEN 5
                        WHEN 'sábado' THEN 6
                        WHEN 'domingo' THEN 7
                        ELSE 8
                    END
            """), {"camp": id_campania})

            contenido_programado = result.fetchall()
            # 🔒 Si es campaña Osceola, forzar solo el ID 169
            # if campaign_key == "botanica_san_gregorio":
            #     contenido_programado = [fila for fila in contenido_programado if fila[0] == 512]


        # Obtener solo el contenido del día actual
        idx_actual = datetime.today().weekday()
        dia_actual = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"][idx_actual]

        contenido_dias = [f for f in contenido_programado if f[1].strip().lower() == dia_actual]
        
        # Obtener los días sábado, domingo y lunes
        # dias_deseados = {"sábado", "domingo", "lunes"}
        # orden_personalizado = {"sábado": 0, "domingo": 1, "lunes": 2}

        # contenido_dias = sorted(
        #     [f for f in contenido_programado if f[1].strip().lower() in dias_deseados],
        #     key=lambda x: orden_personalizado.get(x[1].strip().lower(), 999)
        # )


        if not contenido_dias:
            print(f"⚠️ No hay contenido para el día actual: {dia_actual}")
            exit()


        if not contenido_dias:
            print(f"⚠️ No hay contenido para los días: {dia_actual}")
            exit()
        
        fecha_base = datetime.today().date()
        fechas_por_id_contenido = {
            fila[0]: fecha_base + timedelta(days=idx) for idx, fila in enumerate(contenido_dias)
        }
                
        result = session.execute(sql_text("""
            SELECT main_cta
            FROM annual_schedule
            WHERE id_campaign = :id AND LOWER(TRIM(month)) = :mes
        """), {"id": id_campania, "mes": nombre_mes_actual.lower().strip()})

        row = result.fetchone()
        if not row:
            print(f"⚠️ No hay CTA en annual_schedule para el mes {mes_actual} y campaña ID {id_campania}")
            main_cta_final = "No aplica"
        else:
            main_cta_raw = row[0]
            cta_list = [cta.strip() for cta in (main_cta_raw or "").split("\n") if cta.strip()]
            cta_elegidos = random.sample(cta_list, k=min(len(cta_list), random.choice([1, 2])))
            main_cta_final = ", ".join(cta_elegidos)
        
        # Asignar fechas secuenciales a cada fila de contenido
        for fila in contenido_dias:
            id_contenido, dia, canal, servicio, tipo, descripcion, hashtags, sonido = fila

            print(f"\n🎬 Generando video para:{campaign_key}")
            print(f"📆 Día: {dia} | 📲 Canal: {canal}")
            print(f"🧼 Servicio: {servicio} | 🧠 Tipo: {tipo}")
            print(f"📝 Descripción: {descripcion}")
            print(f"🏷️ Hashtags: {hashtags}")
            print(f"🎵 Sonido: {sonido}")

            gpt = GPT({
                "service": servicio,
                "campaign": campaign_key,
                "lang": lang.lower(),
                "tipo": tipo,
                "descripcion": descripcion,
                "hashtags": hashtags,
                "sonido": sonido,
                "canal": canal,
                "dia": dia,
                "hashtags": hashtags
            })

            if campaign_key == "osceola_fence_corporation":
                theme = gpt.theme_osceola()
                data = {
                    "Text": gpt.copy_osceola(theme, 100),
                    "Document title": gpt.document_title_osceola(theme),
                    "Youtube Video Title": gpt.youtube_video_title_osceola(theme, 40),
                    "Youtube Video Tags": gpt.youtube_video_tags_osceola(theme),
                    "First Comment Text": gpt.first_comment_osceola(theme, 50),
                    "TikTok Title": gpt.tikTok_title_osceola(theme, 50),
                }

            elif campaign_key == "quick_cleaning":
                theme = gpt.theme_quick_cleaning()
                data = {
                    "Text": gpt.copy_quick_cleaning(theme, 100),
                    "Document title": gpt.document_title_quick_cleaning(theme),
                    "Youtube Video Title": gpt.youtube_video_title_quick_cleaning(theme),
                    "Youtube Video Tags": gpt.youtube_video_tags_quick_cleaning(theme),
                    "First Comment Text": gpt.first_comment_quick_cleaning(theme, 50),
                    "TikTok Title": gpt.tikTok_title_quick_cleaning(theme, 50),
                }

            elif campaign_key == "elite_chicago_spa":
                theme = gpt.theme_elite_spa()
                data = {
                    "Text": gpt.copy_elite_spa(theme, 100),
                    "Document title": gpt.document_title_elite_spa(theme),
                    "Youtube Video Title": gpt.youtube_video_title_elite_spa(theme),
                    "Youtube Video Tags": gpt.youtube_video_tags_elite_spa(theme),
                    "First Comment Text": gpt.first_comment_elite_spa(theme, 50),
                    "TikTok Title": gpt.tikTok_title_elite_spa(theme, 50),
                }

            elif campaign_key == "lopez_y_lopez_abogados":
                theme = gpt.theme_lopez_abogados()
                data = {
                    "Text": gpt.copy_lopez_abogados(theme, 100),
                    "Document title": gpt.document_title_lopez_abogados(theme),
                    "Youtube Video Title": gpt.youtube_video_title_lopez_abogados(theme),
                    "Youtube Video Tags": gpt.youtube_video_tags_lopez_abogados(theme),
                    "First Comment Text": gpt.first_comment_lopez_abogados(theme, 50),
                    "TikTok Title": gpt.tikTok_title_lopez_abogados(theme, 50),
                }

            elif campaign_key.startswith("botánica"):
                theme = gpt.theme_botanica()
                data = {
                    "Text": gpt.copy_botanica(theme, 100),
                    "Document title": gpt.document_title_botanica(theme),
                    "Youtube Video Title": gpt.youtube_video_title_botanica(theme, 50),
                    "Youtube Video Tags": gpt.youtube_video_tags_botanica(theme),
                    "First Comment Text": gpt.first_comment_botanica(theme, 50),
                    "TikTok Title": gpt.tikTok_title_botanica(theme, 50),
                }

            elif campaign_key.startswith("botanica"):
                theme = gpt.theme_botanica()
                data = {
                    "Text": gpt.copy_botanica(theme, 100),
                    "Document title": gpt.document_title_botanica(theme),
                    "Youtube Video Title": gpt.youtube_video_title_botanica(theme, 50),
                    "Youtube Video Tags": gpt.youtube_video_tags_botanica(theme),
                    "First Comment Text": gpt.first_comment_botanica(theme, 50),
                    "TikTok Title": gpt.tikTok_title_botanica(theme, 50),
                }

            elif campaign_key.startswith("amarres_chicago"):
                theme = gpt.theme_botanica()
                data = {
                    "Text": gpt.copy_botanica(theme, 100),
                    "Document title": gpt.document_title_botanica(theme),
                    "Youtube Video Title": gpt.youtube_video_title_botanica(theme, 50),
                    "Youtube Video Tags": gpt.youtube_video_tags_botanica(theme),
                    "First Comment Text": gpt.first_comment_botanica(theme, 50),
                    "TikTok Title": gpt.tikTok_title_botanica(theme, 50),
                }
                
            elif campaign_key.startswith("spa312"):
                theme = gpt.theme_botanica()
                data = {
                    "Text": gpt.copy_spa312(theme, 100),
                    "Document title": gpt.document_title_spa312(theme),
                    "Youtube Video Title": gpt.youtube_video_title_spa312(theme, 50),
                    "Youtube Video Tags": gpt.youtube_video_tags_spa312(theme),
                    "First Comment Text": gpt.first_comment_spa312(theme, 50),
                    "TikTok Title": gpt.tikTok_title_spa312(theme, 50),
                }
                
            elif campaign_key.startswith("elite_frenchies"):
                theme = gpt.theme_botanica()
                data = {
                    "Text": gpt.copy_elite_frenchies(theme, 100),
                    "Document title": gpt.document_title_elite_frenchies(theme),
                    "Youtube Video Title": gpt.youtube_video_title_elite_frenchies(theme, 50),
                    "Youtube Video Tags": gpt.youtube_video_tags_elite_frenchies(theme),
                    "First Comment Text": gpt.first_comment_elite_frenchies(theme, 50),
                    "TikTok Title": gpt.tikTok_title_elite_frenchies(theme, 50),
                }

            else:
                print(f"⚠️ Campaña '{campaign_key}' no tiene métodos aún.")
                continue

            print("📥 Texto (Text):", data["Text"])
            print("📄 Título:", data["Document title"])
            plataforma = random.choice(["youtube shorts", "instagram reels", "tiktok"])

            try:
                print("🚀 Ejecutando bot 2...")

                args = [
                    str(data.get("Text", "")),            # tema
                    str(plataforma),                      # plataforma
                    str(descripcion),                     # descripcion
                    str(campaign_key),                    # campaign_key
                    str(lang),                            # language
                    str(canal),                           # canal
                    str(tipo),                            # tipo
                    str(sonido),                          # sonido
                    str(main_cta_final),                  # main_cta
                    str(servicio),                        # servicio
                    str(id_contenido),
                    str(hashtags)
                ]

                print("📦 Argumentos pasados a subprocess:", args)

                subprocess.run(
                    ["python", "main.py"] + args,
                    cwd=r"C:\Users\Programador2\Documents\Antonio Barreto\Communitys\bot_creacion_reels",
                    check=True,
                    timeout=1800
                )

            except subprocess.CalledProcessError as e:
                print(f"❌ Error al ejecutar bot 2: {e}")
                continue

        post_generados += 1

        print("📂 Generando Excel desde base de datos...")
        with SessionLocal() as session:
            query = sql_text("""
                SELECT v.id, v.description, v.platform_video, v.url_drive, c.canal
                FROM videos v
                JOIN contenido_semanal c ON v.id_contenido = c.id
                WHERE v.campaign = :camp AND v.upload_drive = true AND v.upload_excel = false
                ORDER BY v.id
            """)
            resultados = session.execute(query, {"camp": campaign_key}).fetchall()

            for i, fila_db in enumerate(resultados):
                video_id, descripcion, plataforma, url, canal_db = fila_db
                fecha_reel = fecha_actual

                 # Obtener día actual en inglés y español desde la lista de tuplas
                weekday_index = fecha_reel.weekday()  # 0 = lunes, ..., 6 = domingo
                dia_actual = dias_semana[weekday_index]

                # Consulta con IN para aceptar ambos idiomas
                horario_query = sql_text("""
                    SELECT hour FROM schedules 
                    WHERE platform_video = :plat AND day IN :dias
                """).bindparams(bindparam("dias", expanding=True))

                schedules = session.execute(horario_query, {"plat": plataforma, "dias": [dia_actual]}).fetchall()
                ahora = datetime.now().time()

                def generar_hora_posterior():
                    """Genera una hora entre 1h y 1h30 después de ahora, sin pasar de las 23:59"""
                    base = datetime.combine(datetime.today(), ahora)
                    nueva_hora = base + timedelta(minutes=random.randint(60, 90))
                    hora_generada = nueva_hora.time()
                    if hora_generada > time(23, 59):
                        return time(23, 59)
                    if hora_generada < time(17, 0):
                        return time(17, random.randint(0, 59))
                    return hora_generada

                if schedules:
                    horarios_validos = [h[0] for h in schedules if h[0] >= max(time(17, 0), ahora)]
                    if horarios_validos:
                        horario_elegido = random.choice(horarios_validos)
                    else:
                        print("⚠️ No hay horarios válidos en plataforma/día. Buscando en toda la tabla.")
                        query_general = sql_text("SELECT hour FROM schedules")
                        horarios_generales = session.execute(query_general).fetchall()
                        horarios_generales_posteriores = [h[0] for h in horarios_generales if h[0] >= max(time(17, 0), ahora)]
                        if horarios_generales_posteriores:
                            horario_elegido = random.choice(horarios_generales_posteriores)
                        else:
                            print("⚠️ No hay horarios después de las 17:00. Generando 1h o 1h30 después de ahora.")
                            horario_elegido = generar_hora_posterior()
                else:
                    print("⚠️ No se encontraron horarios disponibles. Buscando en toda la tabla.")
                    query_general = sql_text("SELECT hour FROM schedules")
                    horarios_generales = session.execute(query_general).fetchall()
                    horarios_generales_posteriores = [h[0] for h in horarios_generales if h[0] >= max(time(17, 0), ahora)]
                    if horarios_generales_posteriores:
                        horario_elegido = random.choice(horarios_generales_posteriores)
                    else:
                        print("⚠️ No hay horarios después de las 17:00. Generando 1h o 1h30 después de ahora.")
                        horario_elegido = generar_hora_posterior()

                hora_final = horario_elegido.strftime("%H:%M:%S")

                gpt = GPT({"service": "comentario", "campaign": campaign_key, "lang": lang.lower()})
                comentario = data["First Comment Text"]

                for row in range(2, ws.max_row + 1):
                    canal_limpio = canal_db.strip().lower()

                    # Si ya tiene texto, pasamos a la siguiente fila
                    if ws.cell(row=row, column=idx_text).value:
                        continue

                    # Si es Story, no agregamos texto pero sí reservamos la fila
                    if "story" in canal_limpio:
                        ws.cell(row=row, column=idx_text).value = ""
                    else:
                        ws.cell(row=row, column=idx_text).value = descripcion
                        fecha_reel = fechas_por_id_contenido.get(video_id, datetime.today().date())
                        ws.cell(row=row, column=idx_date).value = fecha_reel.strftime("%Y-%m-%d")
                        ws.cell(row=row, column=idx_title).value = data["Document title"]
                        ws.cell(row=row, column=idx_yt_title).value = data["Youtube Video Title"]
                        ws.cell(row=row, column=idx_yt_tags).value = str(post_generados + 1)
                        ws.cell(row=row, column=idx_fb_title).value = str(post_generados + 1)
                        ws.cell(row=row, column=idx_tiktok).value = str(post_generados + 1)
                        
                        if campaign_key == "quick_cleaning":
                            comentario = gpt.first_comment_quick_cleaning(theme, 90)
                        elif campaign_key == "elite_chicago_spa":
                            comentario = gpt.first_comment_elite_spa(theme, 90)
                        elif campaign_key == "lopez_y_lopez_abogados":
                            comentario = gpt.first_comment_lopez_abogados(theme, 90)
                        elif campaign_key.startswith("botanica") or campaign_key.startswith("amarres_chicago"):
                            comentario = gpt.first_comment_botanica(theme, 90)
                        elif campaign_key == "osceola_fence_corporation":
                            comentario = gpt.first_comment_osceola(theme, 90)
                        else:
                            comentario = gpt.comment_from_title(theme, campaign_name=campaign_key, characters=90)

                        # Escribir el comentario generado a la celda
                        ws.cell(row=row, column=idx_comment).value = comentario
                        ws.cell(row=row, column=idx_picture).value = url
                        ws.cell(row=row, column=idx_time).value = hora_final
                        if idx_brand:
                            ws.cell(row=row, column=idx_brand).value = campaign_name
                        for idx_col in idx_autoincrement.values():
                            ws.cell(row=row, column=idx_col).value = str(post_generados + 1)

                        # === Marcar plataformas con el canal real ===
                        canal_limpio = canal_db.strip().lower()
                        col_idx_map = { (cell.value.strip() if cell.value else ""): idx + 1 for idx, cell in enumerate(ws[1]) }

                        if canal_limpio == "tiktok":
                            if "TikTok" in col_idx_map:
                                ws.cell(row=row, column=col_idx_map["TikTok"]).value = "VERDADERO"
                        elif canal_limpio == "ig reels":
                            if "Instagram" in col_idx_map:
                                ws.cell(row=row, column=col_idx_map["Instagram"]).value = "VERDADERO"
                            if "Instagram Post Type" in col_idx_map:
                                ws.cell(row=row, column=col_idx_map["Instagram Post Type"]).value = "REEL"
                            if "Instagram Show Reel On Feed" in col_idx_map:
                                ws.cell(row=row, column=col_idx_map["Instagram Show Reel On Feed"]).value = "VERDADERO"
                        elif canal_limpio == "ig stories":
                            if "Instagram" in col_idx_map:
                                ws.cell(row=row, column=col_idx_map["Instagram"]).value = "VERDADERO"
                            if "Instagram Post Type" in col_idx_map:
                                ws.cell(row=row, column=col_idx_map["Instagram Post Type"]).value = "Story"
                        elif canal_limpio == "yt shorts":
                            if "Youtube" in col_idx_map:
                                ws.cell(row=row, column=col_idx_map["Youtube"]).value = "VERDADERO"
                                if idx_yt_title:
                                    ws.cell(row=row, column=idx_yt_title).value = data["Youtube Video Title"]
                            if "Youtube Video Type" in col_idx_map:
                                ws.cell(row=row, column=col_idx_map["Youtube Video Type"]).value = "SHORT"
                        elif canal_limpio == "fb reels":
                            if "Facebook" in col_idx_map:
                                ws.cell(row=row, column=col_idx_map["Facebook"]).value = "VERDADERO"
                            if "Facebook Post Type" in col_idx_map:
                                ws.cell(row=row, column=col_idx_map["Facebook Post Type"]).value = "REEL"
                        elif canal_limpio == "fb stories":
                            if "Facebook" in col_idx_map:
                                ws.cell(row=row, column=col_idx_map["Facebook"]).value = "VERDADERO"
                            if "Facebook Post Type" in col_idx_map:
                                ws.cell(row=row, column=col_idx_map["Facebook Post Type"]).value = "Story"

                        print(f"✅ Fila {row} completada.")
                        post_generados += 1
                        break
                    
                    # ✅ Marcar el video como procesado en la base de datos
            with SessionLocal() as update_session:
                update_session.execute(
                    sql_text("""
                        UPDATE videos
                        SET upload_excel = true
                        WHERE upload_excel = false AND campaign = :camp
                    """),
                    {"camp": campaign_key}
                )
                update_session.commit()
                print(f"🔁 Todos los videos de la campaña '{campaign_key}' marcados como 'upload_excel = true'")

            # ✅ Elimina filas vacías (donde la celda 'Text' está vacía menos las que son Story)
            max_row = ws.max_row
            for row in range(max_row, 1, -1):
                instagram_type = ws.cell(row=row, column=col_idx_map.get("Instagram Post Type", 0)).value or ""
                facebook_type = ws.cell(row=row, column=col_idx_map.get("Facebook Post Type", 0)).value or ""

                if (
                    not ws.cell(row=row, column=idx_text).value and
                    "story" not in instagram_type.lower() and
                    "story" not in facebook_type.lower()
                ):
                    ws.delete_rows(row)

            # Guarda Excel normal
            wb.save(excel_path)
            print(f"📊 Excel guardado: {excel_path}")

           # Convierte como CSV duplicando los datos del Excel
            csv_path = excel_path.replace(".xlsx", ".csv")
            with open(csv_path, "w", encoding="utf-8", newline="") as f:
                from csv import writer
                csv_writer = writer(f)

                # Escribir encabezados
                csv_writer.writerow([cell.value for cell in ws[1]])

                # Escribir filas
                for row in ws.iter_rows(min_row=2, values_only=False):  # Cambiado a False para poder acceder a las celdas
                    # Verificar si es Story
                    instagram_type = (row[headers.index("Instagram Post Type")].value or "").lower()
                    facebook_type = (row[headers.index("Facebook Post Type")].value or "").lower()
                    
                    # Si es tipo Story, eliminar el valor de "Text"
                    if "story" in instagram_type or "story" in facebook_type:
                        row[headers.index("Text")].value = ""  # Eliminar el texto en la columna "Text"
                    
                    # Escribir la fila modificada en el CSV
                    csv_writer.writerow([cell.value for cell in row])

            print(f"📄 También guardado como CSV: {csv_path}")
            
            try:
                print("🚀 Ejecutando bot 3 para subir excels a Metricool...")
                subprocess.run(
                    ["python", "main.py", str(csv_path)],
                    cwd=r"C:\Users\Programador2\Documents\Antonio Barreto\Communitys\bot_metricool_reels",
                    check=True
                )
            except subprocess.CalledProcessError as e:
                print(f"❌ Error al ejecutar bot 3: {e}")


    print("\n✅ Todas las campañas han sido procesadas.")
    
if __name__ == '__main__':
    # uvicorn.run("fastApi:app", host="0.0.0.0", port=8000, reload=False, access_log=False)
    run_bot()